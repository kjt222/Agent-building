"""Playwright UI smoke for the ExcelRuntimeEdit structured diff (P12.2.2).

Scripted adapter calls ExcelRuntimeEdit with three ops (set_cell, set_formula,
refresh_calculation). The UI should render a structured ops card with the
three rows. Clicking Accept actually mutates the workbook on disk.

Skips silently when win32com / Excel is unavailable so the runner stays
portable.
"""

from __future__ import annotations

import argparse
import json
import os
import socket
import subprocess
import sys
import time
import urllib.request
from contextlib import closing
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

RESULTS = ROOT / "tests" / "results" / "p12_excel_diff_preview_ui"


def _ts() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def find_free_port() -> int:
    with closing(socket.socket(socket.AF_INET, socket.SOCK_STREAM)) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def wait_for_server(base_url: str, timeout_s: float = 30.0) -> None:
    deadline = time.time() + timeout_s
    last = ""
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(f"{base_url}/api/agent_runtime", timeout=3) as resp:
                if 200 <= resp.status < 300:
                    return
        except Exception as exc:
            last = str(exc)
        time.sleep(0.5)
    raise RuntimeError(f"server did not start in {timeout_s}s: {last}")


def _build_fixture(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Q1"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Sales"
    ws["B2"] = 1000
    ws["A3"] = "Tax"
    ws["B3"] = 0
    wb.save(path)


def _cell_value(path: Path, sheet: str, cell: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    return wb[sheet][cell].value


def _write_config(run_dir: Path) -> Path:
    cfg = run_dir / "config"
    cfg.mkdir(parents=True, exist_ok=True)
    (cfg / "app.yaml").write_text(
        """
active_profile: smoke
profiles:
  smoke: {}
active_kbs: []
knowledge_bases: []
runtime:
  mode: inline
  access_mode: restricted
  monitor:
    enabled: false
    wake_on_task_complete: false
    heartbeat_seconds: 30
""".strip(),
        encoding="utf-8",
    )
    (cfg / "models.yaml").write_text(
        """
profiles:
  smoke:
    llm:
      active: provider
      providers:
        provider:
          type: openai
          model: gpt-5-mini
          api_key_ref: smoke.llm.provider
""".strip(),
        encoding="utf-8",
    )
    return cfg


def _build_fake_adapter_module(run_dir: Path, target_file: Path) -> Path:
    sc_dir = run_dir / "sitepatch"
    sc_dir.mkdir(parents=True, exist_ok=True)
    (sc_dir / "sitecustomize.py").write_text(
        f"""
from agent.core.loop import TextDelta, ToolUseDelta, TurnEnd


_TARGET = r"{str(target_file)}"


class _ExcelRuntimeAdapter:
    def __init__(self, model, api_key, base_url=None):
        self.calls = 0

    async def stream(self, messages, tools, system=None, **options):
        self.calls += 1
        if self.calls == 1:
            yield ToolUseDelta(
                id='call-xr-1',
                name='ExcelRuntimeEdit',
                input_partial={{
                    'path': _TARGET,
                    'ops': [
                        {{
                            'op': 'set_cell',
                            'sheet': 'Q1',
                            'cell': 'B3',
                            'value': 100,
                        }},
                        {{
                            'op': 'set_formula',
                            'sheet': 'Q1',
                            'cell': 'B4',
                            'formula': '=B2+B3',
                        }},
                        {{'op': 'refresh_calculation'}},
                    ],
                    'save': True,
                    'refresh_calculation': True,
                }},
            )
            yield TurnEnd(stop_reason='tool_use', usage={{'total_tokens': 1}})
            return
        yield TextDelta(text='done.')
        yield TurnEnd(stop_reason='end_turn', usage={{'total_tokens': 1}})


def _install():
    import agent.models.openai_responses_adapter as mod
    mod.OpenAIResponsesAdapter = _ExcelRuntimeAdapter
    import agent.ui.server as srv
    srv.resolve_api_key = lambda **_: 'smoke-key'


_install()
""".strip(),
        encoding="utf-8",
    )
    return sc_dir


def _skip_with_reason(reason: str) -> int:
    summary = {"passed": True, "skipped": True, "reason": reason, "checks": {}}
    out_dir = RESULTS / _ts()
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def run(_args: argparse.Namespace) -> int:
    try:
        import win32com.client  # noqa: F401
    except Exception:
        return _skip_with_reason("win32com not available")

    # Probe Excel — refuse to spin up the UI if the COM server can't start.
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Quit()
    except Exception as exc:
        return _skip_with_reason(f"Excel.Application unavailable: {exc}")

    from playwright.sync_api import sync_playwright

    run_dir = RESULTS / _ts()
    run_dir.mkdir(parents=True, exist_ok=True)
    target_file = run_dir / "book.xlsx"
    _build_fixture(target_file)
    cfg_dir = _write_config(run_dir)
    sc_dir = _build_fake_adapter_module(run_dir, target_file)
    port = find_free_port()
    base_url = f"http://127.0.0.1:{port}"

    env = dict(os.environ)
    env["PYTHONPATH"] = str(sc_dir) + os.pathsep + env.get("PYTHONPATH", "")
    env["AGENT_CONFIG_DIR"] = str(cfg_dir)

    proc = subprocess.Popen(
        [
            sys.executable, "-m", "uvicorn", "--factory",
            "--host", "127.0.0.1", "--port", str(port),
            "--app-dir", str(ROOT), "agent.ui.server:create_app",
        ],
        env=env, cwd=str(ROOT),
        stdout=(run_dir / "server_stdout.txt").open("w", encoding="utf-8", errors="replace"),
        stderr=(run_dir / "server_stderr.txt").open("w", encoding="utf-8", errors="replace"),
        text=True,
    )

    summary: dict[str, Any] = {"passed": False, "errors": [], "checks": {}}
    try:
        wait_for_server(base_url)
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_context(viewport={"width": 1400, "height": 900}).new_page()
            page.goto(base_url, wait_until="domcontentloaded")
            page.wait_for_selector("#chat-send", timeout=10_000)

            page.fill(
                "#chat-input",
                "请修改这个 Excel 表格：把 B3 改成 100，并在 B4 计算 B2+B3 的合计公式。",
            )
            page.click("#chat-send")
            page.wait_for_selector(".diff-preview-card", timeout=15_000)
            page.wait_for_selector(".diff-preview-op", timeout=5_000)
            card_text = page.locator(".diff-preview-card").last.inner_text(timeout=2_000)
            ops_count = page.locator(".diff-preview-op").count()
            page.screenshot(path=str(run_dir / "01_excel_diff_card.png"), full_page=True)

            page.locator(".diff-accept").last.click()
            page.wait_for_selector(".diff-preview-card.is-accepted", timeout=10_000)
            page.wait_for_function(
                "() => Array.from(document.querySelectorAll('.activity-bar')).some(el => ['done','error'].includes(el.dataset.status))",
                timeout=120_000,
            )
            page.screenshot(path=str(run_dir / "02_after_accept.png"), full_page=True)
            browser.close()

        time.sleep(0.5)
        b3_after = _cell_value(target_file, "Q1", "B3")
        b4_after = _cell_value(target_file, "Q1", "B4")
        normalized = card_text.lower()
        summary["card_text"] = card_text
        summary["ops_card_count"] = ops_count
        summary["b3_after"] = b3_after
        summary["b4_after"] = b4_after
        summary["checks"] = {
            "card_has_set_cell_row": "set cell" in normalized,
            "card_has_set_formula_row": "set formula" in normalized,
            "card_has_refresh_row": "refresh" in normalized or "recalculate" in normalized,
            "card_shows_coord_q1_b3": "Q1!B3" in card_text,
            "card_shows_formula_text": "=B2+B3" in card_text,
            "b3_actually_set": b3_after in (100, 100.0, "100"),
            "b4_is_formula_or_sum": (
                isinstance(b4_after, str) and b4_after.startswith("=")
            ) or b4_after in (1100, 1100.0, 1100.00, "=B2+B3"),
        }
        summary["passed"] = all(summary["checks"].values())
        if not summary["passed"]:
            summary["errors"].append("checks failed: " + json.dumps(summary["checks"]))
    except Exception as exc:
        summary["errors"].append(f"{type(exc).__name__}: {exc}")
    finally:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        (run_dir / "summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    print(json.dumps({
        "passed": summary["passed"],
        "errors": summary["errors"],
        "checks": summary.get("checks"),
    }, ensure_ascii=False, indent=2))
    return 0 if summary["passed"] else 1


def main() -> int:
    parser = argparse.ArgumentParser()
    args = parser.parse_args()
    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())
