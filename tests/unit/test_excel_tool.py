from __future__ import annotations

import asyncio
import json
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from agent.core.loop import LoopConfig, LoopContext, PermissionLevel
from agent.tools_v2.excel_tool import ExcelEditTool, ExcelReadTool, excel_toolset


def _ctx() -> LoopContext:
    return LoopContext(config=LoopConfig())


def _workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Revenue"
    ws["B1"] = 10
    ws["A2"] = "Cost"
    ws["B2"] = 4
    hidden = wb.create_sheet("Notes")
    hidden["A1"] = "keep"
    wb.save(path)


def test_excel_read_reports_cells_and_tracks_read_guard(tmp_path):
    target = tmp_path / "sample.xlsx"
    _workbook(target)
    ctx = _ctx()

    result = asyncio.run(
        ExcelReadTool().run(
            {"path": str(target), "sheet": "Summary", "range": "A1:B2"},
            ctx,
        )
    )

    assert result.is_error is False
    data = json.loads(result.content)
    assert data["type"] == "excel_read"
    assert data["sheets"] == ["Summary", "Notes"]
    assert data["inspected_sheets"][0]["cells"][0]["address"] == "A1"
    assert str(target.resolve()) in ctx.scratch["excel_read_files"]


def test_excel_edit_requires_excel_read_first(tmp_path):
    target = tmp_path / "sample.xlsx"
    _workbook(target)

    result = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [{"op": "set_cell", "sheet": "Summary", "cell": "B1", "value": 12}],
            },
            _ctx(),
        )
    )

    assert result.is_error is True
    assert "Call ExcelRead first" in result.content
    assert openpyxl.load_workbook(target)["Summary"]["B1"].value == 10


def test_excel_edit_applies_scoped_change_and_preserves_unrelated_sheet(tmp_path):
    target = tmp_path / "sample.xlsx"
    _workbook(target)
    ctx = _ctx()
    asyncio.run(ExcelReadTool().run({"path": str(target)}, ctx))

    result = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [
                    {"op": "set_cell", "sheet": "Summary", "cell": "B1", "value": 12},
                    {
                        "op": "set_range_style",
                        "sheet": "Summary",
                        "range": "A1:B1",
                        "font_bold": True,
                        "fill_color": "D9EAD3",
                    },
                ],
            },
            ctx,
        )
    )

    assert result.is_error is False
    data = json.loads(result.content)
    assert data["ops_applied"] == 2
    assert data["backup_path"]
    updated = openpyxl.load_workbook(target)
    assert updated["Summary"]["B1"].value == 12
    assert updated["Summary"]["A1"].font.bold is True
    assert updated["Notes"]["A1"].value == "keep"
    assert str(target.resolve()) in ctx.scratch["edited_files"]


def test_excel_edit_rejects_implicit_sheet_and_large_ranges(tmp_path):
    target = tmp_path / "sample.xlsx"
    _workbook(target)
    ctx = _ctx()
    asyncio.run(ExcelReadTool().run({"path": str(target)}, ctx))

    missing_sheet = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [{"op": "set_cell", "cell": "B1", "value": 12}],
            },
            ctx,
        )
    )
    large_range = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [
                    {
                        "op": "set_range_style",
                        "sheet": "Summary",
                        "range": "A1:Z100",
                        "font_bold": True,
                    }
                ],
            },
            ctx,
        )
    )

    assert missing_sheet.is_error is True
    assert "explicit 'sheet'" in missing_sheet.content
    assert large_range.is_error is True
    assert "allow_large_scope" in large_range.content


def test_excel_edit_copies_template_style_and_dimensions(tmp_path):
    target = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    template = wb.active
    template.title = "Template"
    template["A1"] = "Metric"
    template["B1"] = "Value"
    for cell in template["A1:B1"][0]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin", color="FF000000"))
    template.column_dimensions["A"].width = 18
    template.column_dimensions["B"].width = 12
    template.row_dimensions[1].height = 24

    report = wb.create_sheet("Report")
    report["A1"] = "Metric"
    report["B1"] = "Value"
    report["A2"] = "Revenue"
    report["B2"] = 10
    notes = wb.create_sheet("Notes")
    notes["A1"] = "keep"
    wb.save(target)

    ctx = _ctx()
    asyncio.run(
        ExcelReadTool().run(
            {"path": str(target), "sheet": "Template", "range": "A1:B1"},
            ctx,
        )
    )
    result = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [
                    {
                        "op": "copy_range_style",
                        "source_sheet": "Template",
                        "source_range": "A1:B1",
                        "sheet": "Report",
                        "range": "A1:B1",
                    },
                    {
                        "op": "set_cell",
                        "sheet": "Report",
                        "cell": "B2",
                        "value": 12,
                    },
                ],
            },
            ctx,
        )
    )

    assert result.is_error is False
    updated = openpyxl.load_workbook(target)
    src = updated["Template"]
    dst = updated["Report"]
    assert dst["A1"].font.bold == src["A1"].font.bold
    assert dst["A1"].font.color.rgb == src["A1"].font.color.rgb
    assert dst["A1"].fill.fgColor.rgb == src["A1"].fill.fgColor.rgb
    assert dst["A1"].alignment.horizontal == "center"
    assert dst["A1"].border.bottom.style == "thin"
    assert dst.column_dimensions["A"].width == src.column_dimensions["A"].width
    assert dst.row_dimensions[1].height == src.row_dimensions[1].height
    assert dst["B2"].value == 12
    assert updated["Notes"]["A1"].value == "keep"


def test_excel_edit_rejects_mismatched_copy_style_ranges(tmp_path):
    target = tmp_path / "template.xlsx"
    _workbook(target)
    wb = openpyxl.load_workbook(target)
    template = wb.create_sheet("Template")
    template["A1"] = "Header"
    template["A1"].font = copy(wb["Summary"]["A1"].font)
    wb.save(target)
    ctx = _ctx()
    asyncio.run(ExcelReadTool().run({"path": str(target)}, ctx))

    result = asyncio.run(
        ExcelEditTool().run(
            {
                "path": str(target),
                "ops": [
                    {
                        "op": "copy_range_style",
                        "source_sheet": "Template",
                        "source_range": "A1:A1",
                        "sheet": "Summary",
                        "range": "A1:B1",
                    }
                ],
            },
            ctx,
        )
    )

    assert result.is_error is True
    assert "same shape" in result.content


def test_excel_tool_protocol_flags_are_minimal():
    tools = excel_toolset()

    assert set(tools) == {"ExcelRead", "ExcelEdit"}
    assert tools["ExcelRead"].permission_level == PermissionLevel.SAFE
    assert tools["ExcelEdit"].permission_level == PermissionLevel.NEEDS_APPROVAL
    assert tools["ExcelRead"].parallel_safe is True
    assert tools["ExcelEdit"].parallel_safe is False
