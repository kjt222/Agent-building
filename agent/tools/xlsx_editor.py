from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


@dataclass(frozen=True)
class XlsxOpAction:
    index: int
    action: str
    detail: str


@dataclass
class XlsxEditResult:
    set_cells: int = 0
    formula_cells: int = 0
    inserted_columns: int = 0
    deleted_columns: int = 0
    inserted_rows: int = 0
    deleted_rows: int = 0
    filters_set: int = 0
    sorted_ranges: int = 0


def _normalize_formula(value: str) -> str:
    value = value.strip()
    if not value.startswith("="):
        return f"={value}"
    return value


def _is_empty(value: object) -> bool:
    return value is None or value == ""


def _get_sheet(workbook, sheet_name: Optional[str], default_sheet: Optional[str]):
    if sheet_name:
        return workbook[sheet_name]
    if default_sheet:
        return workbook[default_sheet]
    return workbook.active


class XlsxEditor:
    def __init__(self, path: Path) -> None:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl not installed. Install `openpyxl`.") from exc
        if not path.exists():
            raise FileNotFoundError(path)
        self.path = path
        self.workbook = openpyxl.load_workbook(path)

    def classify_ops(self, ops: list[dict], default_sheet: Optional[str] = None) -> list[XlsxOpAction]:
        actions: list[XlsxOpAction] = []
        for index, op in enumerate(ops):
            op_type = op.get("op")
            sheet = _get_sheet(self.workbook, op.get("sheet"), default_sheet)
            if op_type == "set_cell":
                cell_ref = op.get("cell")
                if not cell_ref:
                    raise ValueError("set_cell requires 'cell'")
                cell = sheet[cell_ref]
                if "formula" in op:
                    action = "tool.xlsx_add_formula" if _is_empty(cell.value) else "tool.xlsx_update_formula"
                else:
                    action = "tool.xlsx_set_value"
                detail = f"{sheet.title}!{cell_ref}"
            elif op_type == "fill_formula":
                ref = op.get("range")
                if not ref:
                    raise ValueError("fill_formula requires 'range'")
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                has_values = False
                for row in sheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    for cell in row:
                        if not _is_empty(cell.value):
                            has_values = True
                            break
                    if has_values:
                        break
                action = "tool.xlsx_update_formula" if has_values else "tool.xlsx_add_formula"
                detail = f"{sheet.title}!{ref}"
            elif op_type == "insert_columns":
                action = "tool.xlsx_insert_columns"
                detail = f"{sheet.title}:{op.get('index')}"
            elif op_type == "delete_columns":
                action = "tool.xlsx_delete_columns"
                detail = f"{sheet.title}:{op.get('index')}"
            elif op_type == "insert_rows":
                action = "tool.xlsx_insert_rows"
                detail = f"{sheet.title}:{op.get('index')}"
            elif op_type == "delete_rows":
                action = "tool.xlsx_delete_rows"
                detail = f"{sheet.title}:{op.get('index')}"
            elif op_type == "set_auto_filter":
                action = "tool.xlsx_filter"
                detail = f"{sheet.title}!{op.get('range')}"
            elif op_type == "sort_range":
                action = "tool.xlsx_sort"
                detail = f"{sheet.title}!{op.get('range')}"
            elif op_type == "pivot":
                action = "tool.xlsx_pivot"
                detail = sheet.title
            else:
                raise ValueError(f"Unsupported xlsx op: {op_type}")
            actions.append(XlsxOpAction(index=index, action=action, detail=detail))
        return actions

    def apply_ops(self, ops: list[dict], default_sheet: Optional[str] = None) -> XlsxEditResult:
        result = XlsxEditResult()
        for op in ops:
            op_type = op.get("op")
            sheet = _get_sheet(self.workbook, op.get("sheet"), default_sheet)
            if op_type == "set_cell":
                cell_ref = op.get("cell")
                if not cell_ref:
                    raise ValueError("set_cell requires 'cell'")
                cell = sheet[cell_ref]
                if "formula" in op:
                    cell.value = _normalize_formula(str(op.get("formula", "")))
                    result.formula_cells += 1
                else:
                    cell.value = op.get("value")
                    result.set_cells += 1
            elif op_type == "fill_formula":
                ref = op.get("range")
                if not ref:
                    raise ValueError("fill_formula requires 'range'")
                template = op.get("formula")
                if template is None:
                    raise ValueError("fill_formula requires 'formula'")
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        formula = str(template).format(
                            row=row,
                            col=get_column_letter(col),
                        )
                        sheet.cell(row=row, column=col, value=_normalize_formula(formula))
                        result.formula_cells += 1
            elif op_type == "insert_columns":
                index = int(op.get("index", 1))
                amount = int(op.get("amount", 1))
                sheet.insert_cols(index, amount)
                result.inserted_columns += amount
            elif op_type == "delete_columns":
                index = int(op.get("index", 1))
                amount = int(op.get("amount", 1))
                sheet.delete_cols(index, amount)
                result.deleted_columns += amount
            elif op_type == "insert_rows":
                index = int(op.get("index", 1))
                amount = int(op.get("amount", 1))
                sheet.insert_rows(index, amount)
                result.inserted_rows += amount
            elif op_type == "delete_rows":
                index = int(op.get("index", 1))
                amount = int(op.get("amount", 1))
                sheet.delete_rows(index, amount)
                result.deleted_rows += amount
            elif op_type == "set_auto_filter":
                ref = op.get("range")
                if not ref:
                    raise ValueError("set_auto_filter requires 'range'")
                sheet.auto_filter.ref = ref
                result.filters_set += 1
            elif op_type == "sort_range":
                ref = op.get("range")
                if not ref:
                    raise ValueError("sort_range requires 'range'")
                key = op.get("key")
                header = bool(op.get("header", False))
                ascending = bool(op.get("ascending", True))
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                start_row = min_row + 1 if header else min_row
                if key is None:
                    key_col = min_col
                elif isinstance(key, int):
                    key_col = int(key)
                else:
                    key_col = column_index_from_string(str(key))
                key_index = key_col - min_col
                rows = []
                for row in sheet.iter_rows(
                    min_row=start_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    rows.append([cell.value for cell in row])
                rows.sort(
                    key=lambda values: (values[key_index] is None, values[key_index]),
                    reverse=not ascending,
                )
                row_cursor = start_row
                for values in rows:
                    for col_offset, value in enumerate(values):
                        sheet.cell(row=row_cursor, column=min_col + col_offset, value=value)
                    row_cursor += 1
                result.sorted_ranges += 1
            elif op_type == "pivot":
                raise NotImplementedError("pivot is not supported yet")
            else:
                raise ValueError(f"Unsupported xlsx op: {op_type}")
        return result

    def save(self) -> None:
        self.workbook.save(self.path)
