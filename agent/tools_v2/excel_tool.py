"""Minimal Excel tools for AgentLoop v2.

The tool boundary intentionally mirrors Claude Code's Read/Edit split:
inspect workbook structure first, then apply scoped structured edits. It does
not execute arbitrary Python, VBA, or COM automation scripts.
"""

from __future__ import annotations

import json
import shutil
import time
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from agent.core.loop import LoopContext, PermissionLevel, ToolResultBlock
from agent.tools_v2.primitives import _ToolBase


_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_MAX_DEFAULT_CELLS = 200
_MAX_EDIT_CELLS = 500


def _load_workbook(path: Path):
    import openpyxl

    keep_vba = path.suffix.lower() == ".xlsm"
    return openpyxl.load_workbook(path, keep_vba=keep_vba)


def _resolve_excel_path(raw_path: Any) -> Path:
    if raw_path is None:
        raise ValueError("path is required")
    path = Path(str(raw_path)).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"file not found: {path}")
    if path.is_dir():
        raise IsADirectoryError(str(path))
    if path.suffix.lower() not in _SUPPORTED_SUFFIXES:
        raise ValueError("expected .xlsx or .xlsm workbook")
    return path


def _path_key(path: Path) -> str:
    return str(path.resolve())


def _cell_count(ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def _range_shape(ref: str) -> tuple[int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return max_row - min_row + 1, max_col - min_col + 1


def _range_ref_for_sheet(sheet) -> str:
    if sheet.max_row < 1 or sheet.max_column < 1:
        return "A1:A1"
    return f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def _color_value(color) -> str | None:
    if color is None:
        return None
    value = getattr(color, "rgb", None)
    if isinstance(value, str):
        return value
    value = getattr(color, "indexed", None)
    if value is not None:
        return f"indexed:{value}"
    value = getattr(color, "theme", None)
    if value is not None:
        return f"theme:{value}"
    return None


def _argb(value: Any) -> str:
    text = str(value).strip().lstrip("#")
    if len(text) == 6:
        return f"FF{text}"
    return text


def _cell_payload(cell, *, include_styles: bool) -> dict:
    payload = {
        "address": cell.coordinate,
        "value": cell.value,
    }
    if isinstance(cell.value, str) and cell.value.startswith("="):
        payload["formula"] = cell.value
    if include_styles:
        payload["style"] = {
            "number_format": cell.number_format,
            "font": {
                "name": cell.font.name,
                "size": cell.font.sz,
                "bold": bool(cell.font.bold),
                "italic": bool(cell.font.italic),
                "color": _color_value(cell.font.color),
            },
            "fill": {
                "fill_type": cell.fill.fill_type,
                "color": _color_value(cell.fill.fgColor),
            },
            "alignment": {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": bool(cell.alignment.wrap_text),
            },
        }
    return payload


def _worksheet_payload(sheet, *, ref: str | None, include_styles: bool, max_cells: int) -> dict:
    target_ref = ref or _range_ref_for_sheet(sheet)
    min_col, min_row, max_col, max_row = range_boundaries(target_ref)
    cells: list[dict] = []
    truncated = False
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if cell.value is None and not include_styles:
                continue
            if len(cells) >= max_cells:
                truncated = True
                break
            cells.append(_cell_payload(cell, include_styles=include_styles))
        if truncated:
            break

    return {
        "name": sheet.title,
        "used_range": _range_ref_for_sheet(sheet),
        "inspected_range": target_ref,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": [str(rng) for rng in sheet.merged_cells.ranges],
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter": sheet.auto_filter.ref,
        "row_heights": {
            str(row): sheet.row_dimensions[row].height
            for row in range(min_row, max_row + 1)
            if sheet.row_dimensions[row].height is not None
        },
        "column_widths": {
            get_column_letter(col): sheet.column_dimensions[get_column_letter(col)].width
            for col in range(min_col, max_col + 1)
            if sheet.column_dimensions[get_column_letter(col)].width is not None
        },
        "cells": cells,
        "truncated": truncated,
    }


def _normalize_formula(value: Any) -> str:
    text = str(value or "").strip()
    return text if text.startswith("=") else f"={text}"


def _require_sheet(op: dict) -> str:
    sheet = str(op.get("sheet") or "").strip()
    if not sheet:
        raise ValueError("each ExcelEdit op requires an explicit 'sheet'")
    return sheet


def _validate_edit_scope(op: dict, *, allow_large_scope: bool) -> None:
    op_type = op.get("op")
    if op_type == "set_cell":
        if not op.get("cell"):
            raise ValueError("set_cell requires 'cell'")
        return
    if op_type in {
        "set_range_style",
        "set_number_format",
        "copy_range_style",
        "merge_cells",
        "unmerge_cells",
    }:
        ref = op.get("range")
        if not ref:
            raise ValueError(f"{op_type} requires 'range'")
        if op_type == "set_number_format" and "number_format" not in op:
            raise ValueError("set_number_format requires 'number_format'")
        if op_type == "copy_range_style":
            if not op.get("source_sheet"):
                raise ValueError("copy_range_style requires 'source_sheet'")
            if not op.get("source_range"):
                raise ValueError("copy_range_style requires 'source_range'")
            if _range_shape(str(ref)) != _range_shape(str(op["source_range"])):
                raise ValueError(
                    "copy_range_style source_range and target range must have "
                    "the same shape"
                )
        count = _cell_count(str(ref))
        if count > _MAX_EDIT_CELLS and not allow_large_scope:
            raise ValueError(
                f"{op_type} touches {count} cells; pass allow_large_scope=true "
                "only after explicitly confirming the intended scope"
            )
        return
    if op_type in {"insert_rows", "delete_rows"}:
        if not op.get("index"):
            raise ValueError(f"{op_type} requires 'index'")
        amount = int(op.get("amount", 1))
        if amount < 1:
            raise ValueError(f"{op_type} amount must be positive")
        if amount > 50 and not allow_large_scope:
            raise ValueError(
                f"{op_type} amount {amount} is too large without allow_large_scope"
            )
        return
    if op_type in {"set_column_width", "set_row_height"}:
        if op_type == "set_column_width" and not op.get("column"):
            raise ValueError("set_column_width requires 'column'")
        if op_type == "set_column_width" and "width" not in op:
            raise ValueError("set_column_width requires 'width'")
        if op_type == "set_row_height" and not op.get("row"):
            raise ValueError("set_row_height requires 'row'")
        if op_type == "set_row_height" and "height" not in op:
            raise ValueError("set_row_height requires 'height'")
        return
    raise ValueError(f"unsupported ExcelEdit op: {op_type}")


def _apply_range_style(sheet, ref: str, op: dict) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    touched = 0
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            font_kwargs = {}
            if "font_bold" in op:
                font_kwargs["bold"] = bool(op["font_bold"])
            if "font_italic" in op:
                font_kwargs["italic"] = bool(op["font_italic"])
            if "font_size" in op:
                font_kwargs["sz"] = float(op["font_size"])
            if "font_color" in op:
                font_kwargs["color"] = _argb(op["font_color"])
            if font_kwargs:
                font = copy(cell.font)
                for key, value in font_kwargs.items():
                    setattr(font, key, value)
                cell.font = font
            if "fill_color" in op:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=_argb(op["fill_color"]),
                )
            if "horizontal_alignment" in op or "vertical_alignment" in op:
                cell.alignment = Alignment(
                    horizontal=op.get("horizontal_alignment", cell.alignment.horizontal),
                    vertical=op.get("vertical_alignment", cell.alignment.vertical),
                    wrap_text=cell.alignment.wrap_text,
                )
            if "number_format" in op:
                cell.number_format = str(op["number_format"])
            touched += 1
    return touched


def _copy_range_style(workbook, op: dict) -> int:
    source_sheet_name = str(op["source_sheet"])
    target_sheet_name = str(op["sheet"])
    if source_sheet_name not in workbook.sheetnames:
        raise ValueError(f"source sheet not found: {source_sheet_name}")
    if target_sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet not found: {target_sheet_name}")
    source_sheet = workbook[source_sheet_name]
    target_sheet = workbook[target_sheet_name]
    source_ref = str(op["source_range"])
    target_ref = str(op["range"])
    if _range_shape(source_ref) != _range_shape(target_ref):
        raise ValueError(
            "copy_range_style source_range and target range must have the same shape"
        )

    s_min_col, s_min_row, s_max_col, s_max_row = range_boundaries(source_ref)
    t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(target_ref)
    touched = 0
    for row_offset, source_row in enumerate(range(s_min_row, s_max_row + 1)):
        target_row = t_min_row + row_offset
        for col_offset, source_col in enumerate(range(s_min_col, s_max_col + 1)):
            target_col = t_min_col + col_offset
            source_cell = source_sheet.cell(source_row, source_col)
            target_cell = target_sheet.cell(target_row, target_col)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            touched += 1

    if bool(op.get("include_dimensions", True)):
        for row_offset, source_row in enumerate(range(s_min_row, s_max_row + 1)):
            height = source_sheet.row_dimensions[source_row].height
            if height is not None:
                target_sheet.row_dimensions[t_min_row + row_offset].height = height
        for col_offset, source_col in enumerate(range(s_min_col, s_max_col + 1)):
            source_letter = get_column_letter(source_col)
            target_letter = get_column_letter(t_min_col + col_offset)
            width = source_sheet.column_dimensions[source_letter].width
            if width is not None:
                target_sheet.column_dimensions[target_letter].width = width
    return touched


class ExcelReadTool(_ToolBase):
    name = "ExcelRead"
    description = (
        "Inspect an .xlsx/.xlsm workbook before editing. Returns workbook "
        "sheets, used ranges, merged cells, selected cell values, formulas, and "
        "basic styles. Use this before ExcelEdit."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Workbook path"},
            "sheet": {"type": "string", "description": "Optional sheet name"},
            "range": {"type": "string", "description": "Optional A1 range"},
            "include_styles": {"type": "boolean", "default": True},
            "max_cells": {"type": "integer", "default": _MAX_DEFAULT_CELLS},
        },
        "required": ["path"],
    }
    permission_level = PermissionLevel.SAFE
    parallel_safe = True

    async def run(self, input: dict, ctx: LoopContext) -> ToolResultBlock:
        try:
            path = _resolve_excel_path(input.get("path"))
            wb = _load_workbook(path)
            include_styles = bool(input.get("include_styles", True))
            max_cells = max(1, min(int(input.get("max_cells", _MAX_DEFAULT_CELLS)), 1000))
            sheet_name = input.get("sheet")
            ref = input.get("range")
            sheets = [ws.title for ws in wb.worksheets]
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"sheet not found: {sheet_name}")
                inspected = [
                    _worksheet_payload(
                        wb[str(sheet_name)],
                        ref=str(ref) if ref else None,
                        include_styles=include_styles,
                        max_cells=max_cells,
                    )
                ]
            else:
                inspected = [
                    _worksheet_payload(
                        ws,
                        ref=None,
                        include_styles=include_styles,
                        max_cells=max(1, max_cells // max(1, len(wb.worksheets))),
                    )
                    for ws in wb.worksheets
                ]
            result = {
                "type": "excel_read",
                "path": str(path),
                "sheets": sheets,
                "active_sheet": wb.active.title,
                "inspected_sheets": inspected,
            }
        except Exception as exc:
            return self._err(f"{type(exc).__name__}: {exc}")
        finally:
            try:
                wb.close()  # type: ignore[name-defined]
            except Exception:
                pass

        ctx.scratch.setdefault("excel_read_files", set()).add(_path_key(path))
        return self._ok(json.dumps(result, ensure_ascii=False, indent=2, default=str))


class ExcelEditTool(_ToolBase):
    name = "ExcelEdit"
    description = (
        "Apply scoped structured edits to an .xlsx/.xlsm workbook. The workbook "
        "must have been inspected with ExcelRead in this AgentLoop run. Every op "
        "requires an explicit sheet and cell/range/index; broad edits are "
        "rejected unless allow_large_scope=true. Reuse RenderDocument after edits "
        "when visual layout matters."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Workbook path"},
            "ops": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "op": {
                            "type": "string",
                            "enum": [
                                "set_cell",
                                "set_range_style",
                                "copy_range_style",
                                "set_number_format",
                                "merge_cells",
                                "unmerge_cells",
                                "insert_rows",
                                "delete_rows",
                                "set_column_width",
                                "set_row_height",
                            ],
                        },
                        "sheet": {"type": "string"},
                        "source_sheet": {"type": "string"},
                        "source_range": {"type": "string"},
                        "cell": {"type": "string"},
                        "range": {"type": "string"},
                        "value": {},
                        "formula": {"type": "string"},
                        "number_format": {"type": "string"},
                        "font_bold": {"type": "boolean"},
                        "font_italic": {"type": "boolean"},
                        "font_size": {"type": "number"},
                        "font_color": {"type": "string"},
                        "fill_color": {"type": "string"},
                        "horizontal_alignment": {"type": "string"},
                        "vertical_alignment": {"type": "string"},
                        "index": {"type": "integer"},
                        "amount": {"type": "integer", "default": 1},
                        "column": {"type": "string"},
                        "row": {"type": "integer"},
                        "width": {"type": "number"},
                        "height": {"type": "number"},
                        "include_dimensions": {"type": "boolean", "default": True},
                    },
                    "required": ["op", "sheet"],
                },
            },
            "backup": {"type": "boolean", "default": True},
            "allow_large_scope": {"type": "boolean", "default": False},
        },
        "required": ["path", "ops"],
    }
    permission_level = PermissionLevel.NEEDS_APPROVAL
    parallel_safe = False

    async def run(self, input: dict, ctx: LoopContext) -> ToolResultBlock:
        try:
            path = _resolve_excel_path(input.get("path"))
            if _path_key(path) not in ctx.scratch.setdefault("excel_read_files", set()):
                return self._err(
                    f"workbook was not inspected in this session: {path}. "
                    "Call ExcelRead first."
                )
            ops = input.get("ops")
            if not isinstance(ops, list) or not ops:
                return self._err("ExcelEdit requires a non-empty ops list")
            allow_large_scope = bool(input.get("allow_large_scope", False))
            for op in ops:
                if not isinstance(op, dict):
                    raise ValueError("each op must be an object")
                _require_sheet(op)
                _validate_edit_scope(op, allow_large_scope=allow_large_scope)

            backup_path = None
            if bool(input.get("backup", True)):
                stamp = time.strftime("%Y%m%d%H%M%S")
                backup_path = path.with_name(f"{path.stem}.bak-{stamp}{path.suffix}")
                shutil.copy2(path, backup_path)

            wb = _load_workbook(path)
            touched: list[str] = []
            for op in ops:
                sheet_name = _require_sheet(op)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"sheet not found: {sheet_name}")
                sheet = wb[sheet_name]
                op_type = op["op"]
                if op_type == "set_cell":
                    cell = sheet[str(op["cell"])]
                    if "formula" in op:
                        cell.value = _normalize_formula(op["formula"])
                    else:
                        cell.value = op.get("value")
                    if "number_format" in op:
                        cell.number_format = str(op["number_format"])
                    touched.append(f"{sheet_name}!{cell.coordinate}")
                elif op_type == "set_range_style":
                    count = _apply_range_style(sheet, str(op["range"]), op)
                    touched.append(f"{sheet_name}!{op['range']} ({count} cells)")
                elif op_type == "copy_range_style":
                    count = _copy_range_style(wb, op)
                    touched.append(
                        f"{op['source_sheet']}!{op['source_range']} -> "
                        f"{sheet_name}!{op['range']} ({count} cells)"
                    )
                elif op_type == "set_number_format":
                    count = _apply_range_style(
                        sheet,
                        str(op["range"]),
                        {"number_format": op["number_format"]},
                    )
                    touched.append(f"{sheet_name}!{op['range']} ({count} cells)")
                elif op_type == "merge_cells":
                    sheet.merge_cells(str(op["range"]))
                    touched.append(f"{sheet_name}!merge:{op['range']}")
                elif op_type == "unmerge_cells":
                    ref = str(op["range"])
                    if ref in [str(rng) for rng in sheet.merged_cells.ranges]:
                        sheet.unmerge_cells(ref)
                    touched.append(f"{sheet_name}!unmerge:{ref}")
                elif op_type == "insert_rows":
                    amount = int(op.get("amount", 1))
                    sheet.insert_rows(int(op["index"]), amount)
                    touched.append(f"{sheet_name}!rows@{op['index']}+{amount}")
                elif op_type == "delete_rows":
                    amount = int(op.get("amount", 1))
                    sheet.delete_rows(int(op["index"]), amount)
                    touched.append(f"{sheet_name}!rows@{op['index']}-{amount}")
                elif op_type == "set_column_width":
                    column = str(op["column"]).upper()
                    width = float(op["width"])
                    sheet.column_dimensions[column].width = width
                    touched.append(f"{sheet_name}!col:{column}")
                elif op_type == "set_row_height":
                    row = int(op["row"])
                    height = float(op["height"])
                    sheet.row_dimensions[row].height = height
                    touched.append(f"{sheet_name}!row:{row}")

            wb.save(path)
            result = {
                "type": "excel_edit",
                "path": str(path),
                "backup_path": str(backup_path) if backup_path else None,
                "ops_applied": len(ops),
                "touched": touched,
            }
            ctx.scratch.setdefault("excel_edited_files", set()).add(_path_key(path))
            ctx.scratch.setdefault("edited_files", set()).add(_path_key(path))
            return self._ok(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        except Exception as exc:
            return self._err(f"{type(exc).__name__}: {exc}")
        finally:
            try:
                wb.close()  # type: ignore[name-defined]
            except Exception:
                pass


def excel_toolset() -> dict:
    tools = [ExcelReadTool(), ExcelEditTool()]
    return {tool.name: tool for tool in tools}
